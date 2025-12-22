#!/usr/bin/env python3
"""
Excel Financial Model Generator

Creates professional Excel models like Claude FSI:
- DCF Models
- Comparable Company Analysis
- LBO Models
- Earnings Analysis

Uses: openpyxl, xlsxwriter, yfinance
"""

import yfinance as yf
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from datetime import datetime
import os

# Styles
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MONEY_FORMAT = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
PERCENT_FORMAT = '0.0%'
MULTIPLE_FORMAT = '0.0x'
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def style_header(ws, row, cols):
    """Apply header styling to a row"""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')

def generate_dcf_model(ticker: str, output_path: str = None) -> str:
    """
    Generate a complete DCF model in Excel

    Args:
        ticker: Stock ticker symbol
        output_path: Output file path (optional)

    Returns:
        Path to generated Excel file
    """
    stock = yf.Ticker(ticker)
    info = stock.info
    financials = stock.income_stmt
    balance = stock.balance_sheet
    cashflow = stock.cashflow

    if output_path is None:
        output_path = f"{ticker}_DCF_Model_{datetime.now().strftime('%Y%m%d')}.xlsx"

    wb = Workbook()

    # ========== SUMMARY TAB ==========
    ws_summary = wb.active
    ws_summary.title = "Summary"

    ws_summary['A1'] = f"DCF Valuation Model: {info.get('longName', ticker)}"
    ws_summary['A1'].font = Font(size=16, bold=True)

    ws_summary['A3'] = "Company Information"
    ws_summary['A3'].font = Font(bold=True)

    summary_data = [
        ["Ticker", ticker],
        ["Company Name", info.get('longName', 'N/A')],
        ["Sector", info.get('sector', 'N/A')],
        ["Industry", info.get('industry', 'N/A')],
        ["Current Price", info.get('currentPrice', 0)],
        ["Market Cap", info.get('marketCap', 0)],
        ["Enterprise Value", info.get('enterpriseValue', 0)],
        ["Shares Outstanding", info.get('sharesOutstanding', 0)],
    ]

    for i, (label, value) in enumerate(summary_data, start=4):
        ws_summary[f'A{i}'] = label
        ws_summary[f'B{i}'] = value

    # ========== ASSUMPTIONS TAB ==========
    ws_assumptions = wb.create_sheet("Assumptions")

    ws_assumptions['A1'] = "DCF Assumptions"
    ws_assumptions['A1'].font = Font(size=14, bold=True)

    assumptions = [
        ["Revenue Growth Assumptions", ""],
        ["Year 1 Growth", 0.10],
        ["Year 2 Growth", 0.08],
        ["Year 3 Growth", 0.06],
        ["Year 4 Growth", 0.05],
        ["Year 5 Growth", 0.04],
        ["Terminal Growth", 0.025],
        ["", ""],
        ["Margin Assumptions", ""],
        ["EBITDA Margin", info.get('ebitdaMargins', 0.15)],
        ["Tax Rate", 0.21],
        ["D&A % of Revenue", 0.03],
        ["CapEx % of Revenue", 0.04],
        ["NWC % of Revenue", 0.05],
        ["", ""],
        ["WACC Assumptions", ""],
        ["Risk-Free Rate", 0.045],
        ["Beta", info.get('beta', 1.0)],
        ["Market Risk Premium", 0.055],
        ["Cost of Debt", 0.06],
        ["Tax Rate (Debt)", 0.21],
        ["Debt Weight", 0.30],
        ["Equity Weight", 0.70],
    ]

    for i, (label, value) in enumerate(assumptions, start=3):
        ws_assumptions[f'A{i}'] = label
        ws_assumptions[f'B{i}'] = value
        if isinstance(value, float) and value < 1:
            ws_assumptions[f'B{i}'].number_format = PERCENT_FORMAT

    # ========== INCOME STATEMENT TAB ==========
    ws_income = wb.create_sheet("Income Statement")

    ws_income['A1'] = "Historical Income Statement"
    ws_income['A1'].font = Font(size=14, bold=True)

    if financials is not None and not financials.empty:
        # Headers
        ws_income['A3'] = "Metric"
        col = 2
        for date in financials.columns:
            ws_income.cell(row=3, column=col, value=date.strftime('%Y') if hasattr(date, 'strftime') else str(date))
            col += 1
        style_header(ws_income, 3, col - 1)

        # Data
        row = 4
        key_metrics = ['Total Revenue', 'Gross Profit', 'Operating Income', 'EBITDA', 'Net Income']
        for metric in financials.index:
            if any(key in str(metric) for key in ['Revenue', 'Profit', 'Income', 'EBITDA']):
                ws_income.cell(row=row, column=1, value=str(metric))
                col = 2
                for date in financials.columns:
                    value = financials.loc[metric, date]
                    ws_income.cell(row=row, column=col, value=float(value) if pd.notna(value) else 0)
                    ws_income.cell(row=row, column=col).number_format = MONEY_FORMAT
                    col += 1
                row += 1

    # ========== DCF CALCULATION TAB ==========
    ws_dcf = wb.create_sheet("DCF Model")

    ws_dcf['A1'] = "Discounted Cash Flow Analysis"
    ws_dcf['A1'].font = Font(size=14, bold=True)

    # Get base revenue
    base_revenue = info.get('totalRevenue', 0)

    # Headers
    years = ['Base', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5', 'Terminal']
    for i, year in enumerate(years):
        ws_dcf.cell(row=3, column=i+2, value=year)
    style_header(ws_dcf, 3, len(years) + 1)

    # Revenue projection
    ws_dcf['A4'] = "Revenue"
    growth_rates = [0, 0.10, 0.08, 0.06, 0.05, 0.04, 0.025]
    revenues = [base_revenue]
    for g in growth_rates[1:]:
        revenues.append(revenues[-1] * (1 + g))

    for i, rev in enumerate(revenues):
        ws_dcf.cell(row=4, column=i+2, value=rev)
        ws_dcf.cell(row=4, column=i+2).number_format = MONEY_FORMAT

    # Growth rate row
    ws_dcf['A5'] = "Growth Rate"
    for i, g in enumerate(growth_rates):
        if i > 0:
            ws_dcf.cell(row=5, column=i+2, value=g)
            ws_dcf.cell(row=5, column=i+2).number_format = PERCENT_FORMAT

    # EBITDA
    ebitda_margin = info.get('ebitdaMargins', 0.15) or 0.15
    ws_dcf['A7'] = "EBITDA"
    ws_dcf['A8'] = "EBITDA Margin"
    for i, rev in enumerate(revenues):
        ws_dcf.cell(row=7, column=i+2, value=rev * ebitda_margin)
        ws_dcf.cell(row=7, column=i+2).number_format = MONEY_FORMAT
        ws_dcf.cell(row=8, column=i+2, value=ebitda_margin)
        ws_dcf.cell(row=8, column=i+2).number_format = PERCENT_FORMAT

    # Free Cash Flow
    ws_dcf['A10'] = "Free Cash Flow Calculation"
    ws_dcf['A10'].font = Font(bold=True)

    ws_dcf['A11'] = "EBITDA"
    ws_dcf['A12'] = "Less: D&A"
    ws_dcf['A13'] = "EBIT"
    ws_dcf['A14'] = "Less: Taxes (21%)"
    ws_dcf['A15'] = "EBIAT"
    ws_dcf['A16'] = "Plus: D&A"
    ws_dcf['A17'] = "Less: CapEx"
    ws_dcf['A18'] = "Less: Change in NWC"
    ws_dcf['A19'] = "Free Cash Flow"
    ws_dcf['A19'].font = Font(bold=True)

    # Calculate FCF for each year
    for i in range(1, len(revenues)):
        col = i + 2
        rev = revenues[i]
        ebitda = rev * ebitda_margin
        da = rev * 0.03
        ebit = ebitda - da
        taxes = ebit * 0.21
        ebiat = ebit - taxes
        capex = rev * 0.04
        nwc_change = (revenues[i] - revenues[i-1]) * 0.05 if i > 0 else 0
        fcf = ebiat + da - capex - nwc_change

        ws_dcf.cell(row=11, column=col, value=ebitda).number_format = MONEY_FORMAT
        ws_dcf.cell(row=12, column=col, value=-da).number_format = MONEY_FORMAT
        ws_dcf.cell(row=13, column=col, value=ebit).number_format = MONEY_FORMAT
        ws_dcf.cell(row=14, column=col, value=-taxes).number_format = MONEY_FORMAT
        ws_dcf.cell(row=15, column=col, value=ebiat).number_format = MONEY_FORMAT
        ws_dcf.cell(row=16, column=col, value=da).number_format = MONEY_FORMAT
        ws_dcf.cell(row=17, column=col, value=-capex).number_format = MONEY_FORMAT
        ws_dcf.cell(row=18, column=col, value=-nwc_change).number_format = MONEY_FORMAT
        ws_dcf.cell(row=19, column=col, value=fcf).number_format = MONEY_FORMAT

    # WACC Calculation
    ws_dcf['A21'] = "WACC Calculation"
    ws_dcf['A21'].font = Font(bold=True)

    rf = 0.045
    beta = info.get('beta', 1.0) or 1.0
    mrp = 0.055
    cost_of_equity = rf + beta * mrp
    cost_of_debt = 0.06
    tax_rate = 0.21
    debt_weight = 0.30
    equity_weight = 0.70
    wacc = equity_weight * cost_of_equity + debt_weight * cost_of_debt * (1 - tax_rate)

    wacc_data = [
        ["Risk-Free Rate", rf],
        ["Beta", beta],
        ["Market Risk Premium", mrp],
        ["Cost of Equity", cost_of_equity],
        ["Cost of Debt (pre-tax)", cost_of_debt],
        ["Tax Rate", tax_rate],
        ["After-tax Cost of Debt", cost_of_debt * (1 - tax_rate)],
        ["Debt Weight", debt_weight],
        ["Equity Weight", equity_weight],
        ["WACC", wacc],
    ]

    for i, (label, value) in enumerate(wacc_data, start=22):
        ws_dcf[f'A{i}'] = label
        ws_dcf[f'B{i}'] = value
        ws_dcf[f'B{i}'].number_format = PERCENT_FORMAT if value < 1 else '0.00'

    ws_dcf['A32'].font = Font(bold=True)
    ws_dcf['A32'] = "WACC"
    ws_dcf['B32'] = wacc
    ws_dcf['B32'].number_format = PERCENT_FORMAT
    ws_dcf['B32'].font = Font(bold=True)

    # ========== SENSITIVITY TAB ==========
    ws_sens = wb.create_sheet("Sensitivity")

    ws_sens['A1'] = "Sensitivity Analysis"
    ws_sens['A1'].font = Font(size=14, bold=True)

    ws_sens['A3'] = "WACC \\ Terminal Growth"
    growth_rates_sens = [0.015, 0.020, 0.025, 0.030, 0.035]
    wacc_rates = [0.08, 0.085, 0.09, 0.095, 0.10]

    # Headers
    for i, g in enumerate(growth_rates_sens):
        ws_sens.cell(row=3, column=i+2, value=g)
        ws_sens.cell(row=3, column=i+2).number_format = PERCENT_FORMAT
    style_header(ws_sens, 3, len(growth_rates_sens) + 1)

    # Calculate sensitivity table
    # Use Year 5 FCF for terminal value
    year5_fcf = revenues[5] * ebitda_margin * 0.7  # Simplified FCF

    for i, w in enumerate(wacc_rates):
        ws_sens.cell(row=i+4, column=1, value=w)
        ws_sens.cell(row=i+4, column=1).number_format = PERCENT_FORMAT

        for j, g in enumerate(growth_rates_sens):
            if w > g:
                terminal_value = year5_fcf * (1 + g) / (w - g)
                # Discount back
                ev = terminal_value / ((1 + w) ** 5)
                equity_value = ev - info.get('totalDebt', 0) + info.get('totalCash', 0)
                shares = info.get('sharesOutstanding', 1)
                implied_price = equity_value / shares if shares > 0 else 0

                ws_sens.cell(row=i+4, column=j+2, value=implied_price)
                ws_sens.cell(row=i+4, column=j+2).number_format = '$#,##0.00'
            else:
                ws_sens.cell(row=i+4, column=j+2, value="N/A")

    # Add conditional formatting
    ws_sens.conditional_formatting.add(
        f'B4:F8',
        ColorScaleRule(
            start_type='min', start_color='F8696B',
            mid_type='percentile', mid_value=50, mid_color='FFEB84',
            end_type='max', end_color='63BE7B'
        )
    )

    # Save
    wb.save(output_path)
    return output_path


def generate_comps_model(ticker: str, peers: list = None, output_path: str = None) -> str:
    """
    Generate comparable company analysis in Excel

    Args:
        ticker: Target stock ticker
        peers: List of peer tickers (optional, will auto-detect)
        output_path: Output file path
    """
    stock = yf.Ticker(ticker)
    info = stock.info

    if peers is None:
        # Try to get recommended symbols
        peers = info.get('recommendedSymbols', [])[:8]
        if not peers:
            # Default tech peers
            peers = ['AAPL', 'MSFT', 'GOOGL', 'AMZN', 'META']

    if output_path is None:
        output_path = f"{ticker}_Comps_{datetime.now().strftime('%Y%m%d')}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Trading Comps"

    ws['A1'] = f"Comparable Company Analysis: {info.get('longName', ticker)}"
    ws['A1'].font = Font(size=16, bold=True)

    # Headers
    headers = [
        'Company', 'Ticker', 'Price', 'Market Cap', 'EV',
        'EV/Revenue', 'EV/EBITDA', 'P/E', 'P/B',
        'Gross Margin', 'EBITDA Margin', 'Net Margin',
        'Revenue Growth', 'ROE'
    ]

    for i, h in enumerate(headers, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, len(headers))

    # Gather data
    all_tickers = [ticker] + [p for p in peers if p != ticker]
    row = 4

    for t in all_tickers:
        try:
            s = yf.Ticker(t)
            i = s.info

            ws.cell(row=row, column=1, value=i.get('longName', t)[:30])
            ws.cell(row=row, column=2, value=t)
            ws.cell(row=row, column=3, value=i.get('currentPrice', 0))
            ws.cell(row=row, column=4, value=i.get('marketCap', 0))
            ws.cell(row=row, column=5, value=i.get('enterpriseValue', 0))
            ws.cell(row=row, column=6, value=i.get('enterpriseToRevenue', 0))
            ws.cell(row=row, column=7, value=i.get('enterpriseToEbitda', 0))
            ws.cell(row=row, column=8, value=i.get('trailingPE', 0))
            ws.cell(row=row, column=9, value=i.get('priceToBook', 0))
            ws.cell(row=row, column=10, value=i.get('grossMargins', 0))
            ws.cell(row=row, column=11, value=i.get('ebitdaMargins', 0))
            ws.cell(row=row, column=12, value=i.get('profitMargins', 0))
            ws.cell(row=row, column=13, value=i.get('revenueGrowth', 0))
            ws.cell(row=row, column=14, value=i.get('returnOnEquity', 0))

            # Format
            ws.cell(row=row, column=3).number_format = '$#,##0.00'
            ws.cell(row=row, column=4).number_format = MONEY_FORMAT
            ws.cell(row=row, column=5).number_format = MONEY_FORMAT
            ws.cell(row=row, column=6).number_format = MULTIPLE_FORMAT
            ws.cell(row=row, column=7).number_format = MULTIPLE_FORMAT
            ws.cell(row=row, column=8).number_format = MULTIPLE_FORMAT
            ws.cell(row=row, column=9).number_format = MULTIPLE_FORMAT
            for col in range(10, 15):
                ws.cell(row=row, column=col).number_format = PERCENT_FORMAT

            # Highlight target
            if t == ticker:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            row += 1
        except Exception as e:
            print(f"Error fetching {t}: {e}")
            continue

    # Statistics row
    row += 1
    ws.cell(row=row, column=1, value="Mean")
    ws.cell(row=row, column=1).font = Font(bold=True)

    row += 1
    ws.cell(row=row, column=1, value="Median")
    ws.cell(row=row, column=1).font = Font(bold=True)

    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=3, column=col).column_letter].width = 15

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 3:
        print("Usage: python excel_generator.py [dcf|comps] TICKER [peers]")
        print("  dcf AAPL           - Generate DCF model for AAPL")
        print("  comps AAPL MSFT,GOOGL - Generate comps for AAPL vs peers")
        sys.exit(1)

    model_type = sys.argv[1]
    ticker = sys.argv[2]

    if model_type == "dcf":
        path = generate_dcf_model(ticker)
        print(f"DCF model generated: {path}")

    elif model_type == "comps":
        peers = sys.argv[3].split(',') if len(sys.argv) > 3 else None
        path = generate_comps_model(ticker, peers)
        print(f"Comps model generated: {path}")

    else:
        print(f"Unknown model type: {model_type}")
